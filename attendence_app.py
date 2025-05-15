import streamlit as st
import pandas as pd
import os
import time
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Set page title and configuration
st.set_page_config(
    page_title="Vistotech Attendance System",
    page_icon="⏰",
    layout="wide"
)

# Add simple styling for the header
st.markdown("""
    <style>
    .title-container {
        display: flex;
        align-items: center;
        background-color: #f8f9fa;
        padding: 10px;
        border-radius: 5px;
        margin-bottom: 10px;
    }
    .company-name {
        color: #333;
        font-size: 24px;
        margin-left: 10px;
    }
    </style>
""", unsafe_allow_html=True)

# Define file path for Excel file
EXCEL_FILE = "attendence_data.xlsx"

# Initialize the Excel file if it doesn't exist
def initialize_excel():
    try:
        # Create a new DataFrame with the required columns
        df = pd.DataFrame(columns=[
            'Employee ID', 
            'Employee Name', 
            'Date', 
            'Punch In Time', 
            'Punch Out Time', 
            'Work Hours',
            'Status',
            'Is Late'  # Field to track lateness
        ])
        
        # Make sure the file doesn't exist before creating a new one
        if os.path.exists(EXCEL_FILE):
            try:
                # Try to load existing data first
                existing_df = pd.read_excel(EXCEL_FILE)
                
                # Check if we need to update the schema from old format to new format
                if 'Break 1 Start' in existing_df.columns:
                    # We're migrating from the old format with breaks to the new format
                    # Create a new dataframe with the new schema
                    new_df = pd.DataFrame(columns=[
                        'Employee ID', 
                        'Employee Name', 
                        'Date', 
                        'Punch In Time', 
                        'Punch Out Time', 
                        'Work Hours',
                        'Status',
                        'Is Late'
                    ])
                    
                    # Copy over the existing data that matches the new schema
                    for col in new_df.columns:
                        if col in existing_df.columns:
                            new_df[col] = existing_df[col]
                    
                    # Calculate Is Late for existing records
                    if 'Punch In Time' in existing_df.columns:
                        def check_if_late(row):
                            if pd.notna(row['Punch In Time']):
                                try:
                                    punch_in_time = datetime.strptime(row['Punch In Time'], '%H:%M:%S').time()
                                    cutoff_time = datetime.strptime("10:15:00", '%H:%M:%S').time()
                                    return punch_in_time > cutoff_time
                                except:
                                    return False
                            return False
                        
                        new_df['Is Late'] = existing_df.apply(check_if_late, axis=1)
                    
                    # Save the updated schema
                    df = new_df
                else:
                    # Use existing data if it has the correct schema
                    df = existing_df
            except Exception as e:
                # If there's an error reading the file, it might be corrupted
                # Delete it and create a new one
                st.warning(f"Recreating Excel file due to error: {e}")
                os.remove(EXCEL_FILE)
        
        # Save the DataFrame to Excel
        df.to_excel(EXCEL_FILE, index=False)
        
        # Verify file was created
        if not os.path.exists(EXCEL_FILE):
            st.error("Failed to create Excel file")
        
        return df
    except Exception as e:
        st.error(f"Error initializing Excel file: {e}")
        # Create a minimal dataframe to return
        return pd.DataFrame(columns=[
            'Employee ID', 
            'Employee Name', 
            'Date', 
            'Punch In Time', 
            'Punch Out Time', 
            'Work Hours',
            'Status',
            'Is Late'
        ])

# Function to load employee data
def load_data():
    try:
        if os.path.exists(EXCEL_FILE):
            df = pd.read_excel(EXCEL_FILE)
            return df
        else:
            return initialize_excel()
    except Exception as e:
        st.error(f"Error loading data: {e}")
        return initialize_excel()

# Function to save data
def save_data(df):
    try:
        # Make sure the dataframe is not empty
        if df is None or df.empty:
            df = pd.DataFrame(columns=[
                'Employee ID', 
                'Employee Name', 
                'Date', 
                'Punch In Time', 
                'Punch Out Time', 
                'Work Hours',
                'Status',
                'Is Late'
            ])
        
        # Save to Excel
        df.to_excel(EXCEL_FILE, index=False)
        
        # Verify file was created and has content
        if os.path.exists(EXCEL_FILE) and os.path.getsize(EXCEL_FILE) > 0:
            return True
        else:
            st.error("Excel file was not created or is empty")
            return False
    except Exception as e:
        st.error(f"Error saving data: {e}")
        return False

# Function to check if employee has already punched in today
def check_existing_punch_in(emp_id, today, emp_name=None):
    df = load_data()
    
    # Convert employee ID to string to ensure comparison works correctly
    emp_id = str(emp_id).strip()
    
    # Filter data for today and this employee ID
    today_records = df[df['Date'] == today]
    
    # Convert all Employee ID values to strings for consistent comparison
    today_records = today_records[today_records['Employee ID'].astype(str).str.strip() == emp_id]
    
    # If name is provided, further filter by name
    if emp_name is not None and len(today_records) > 0:
        # Try exact match first
        name_match = today_records[today_records['Employee Name'].str.lower() == emp_name.lower()]
        
        # If no exact match, return all records for this employee ID
        if len(name_match) > 0:
            today_records = name_match
    
    # Log information (only visible in development)
    if not today_records.empty:
        # Check if there's any record without punch out time
        incomplete_records = today_records[today_records['Punch Out Time'].isna()]
        if not incomplete_records.empty:
            return True, incomplete_records.index[0]
    else:
        # No records found for this employee on this date
        pass
    
    return False, None

# Function to calculate hours worked
def calculate_hours(punch_in, punch_out):
    if pd.isna(punch_out) or pd.isna(punch_in):
        return None
    
    # Convert to datetime objects if they are strings
    def convert_to_time(time_value):
        if pd.isna(time_value):
            return None
        if isinstance(time_value, str):
            return datetime.strptime(time_value, '%H:%M:%S').time()
        return time_value
    
    punch_in = convert_to_time(punch_in)
    punch_out = convert_to_time(punch_out)
    
    # Calculate the total work time
    in_dt = datetime.combine(datetime.today().date(), punch_in)
    out_dt = datetime.combine(datetime.today().date(), punch_out)
    
    # If punch out is before punch in, assume it's the next day
    if out_dt < in_dt:
        out_dt = datetime.combine(datetime.today().date() + pd.Timedelta(days=1), punch_out)
    
    total_time = out_dt - in_dt
    work_hours = total_time.total_seconds() / 3600
    
    return round(work_hours, 2)

# Function to apply colors to Excel cells based on lateness
def apply_excel_formatting():
    try:
        if os.path.exists(EXCEL_FILE):
            # Load the workbook
            wb = load_workbook(EXCEL_FILE)
            sheet = wb.active
            
            # Get late column index (column H, index 7)
            late_col_idx = 7
            
            # Start from row 2 (skip headers)
            for row_idx in range(2, sheet.max_row + 1):
                is_late_cell = sheet.cell(row=row_idx, column=late_col_idx + 1)
                
                # Check if the Is Late field is True
                if is_late_cell.value == True:
                    # Apply red background to the entire row
                    for col_idx in range(1, sheet.max_column + 1):
                        sheet.cell(row=row_idx, column=col_idx).fill = PatternFill(
                            start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                        )
                elif is_late_cell.value == False:
                    # Apply green background to the entire row
                    for col_idx in range(1, sheet.max_column + 1):
                        sheet.cell(row=row_idx, column=col_idx).fill = PatternFill(
                            start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"
                        )
            
            # Save the workbook
            wb.save(EXCEL_FILE)
            return True
    except Exception as e:
        st.error(f"Error applying Excel formatting: {e}")
    return False

# Main application
def main():
    # Display company logo at the top
    col1, col2 = st.columns([1, 3])
    
    # Logo on the left
    if os.path.exists("vistotech_logo.png"):
        col1.image("vistotech_logo.png", width=180)
    
    # Title and subtitle on the right
    col2.title("Vistotech Global Services")
    col2.markdown("**Attendance Management System**")
    
    # Initialize Excel file if it doesn't exist
    if not os.path.exists(EXCEL_FILE):
        initialize_excel()
    
    # Auto-refresh feature for real-time updates
    auto_refresh = st.sidebar.checkbox("Enable Auto-Refresh", value=True)
    refresh_interval = st.sidebar.slider("Refresh Interval (seconds)", 
                                       min_value=10, 
                                       max_value=300, 
                                       value=60)
    
    if auto_refresh:
        # Add auto-refresh functionality using HTML/JavaScript
        st.markdown(
            f"""
            <script>
                setTimeout(function() {{
                    window.location.reload();
                }}, {refresh_interval * 1000});
            </script>
            """,
            unsafe_allow_html=True
        )
        st.sidebar.info(f"Page will auto-refresh every {refresh_interval} seconds")
        
        # Show current timestamp
        st.sidebar.write(f"Last refresh: {datetime.now().strftime('%H:%M:%S')}")
    
    # Sidebar for navigation with separate punch in/out options
    st.sidebar.header("Vistotech Navigation")
    page = st.sidebar.selectbox("Choose a page", ["Punch In", "Punch Out", "View Reports", "Admin Panel"])
    
    if page == "Punch In":
        punch_in_page()
    elif page == "Punch Out":
        punch_out_page()
    elif page == "View Reports":
        # Add password protection to View Reports
        admin_password = get_admin_password()
        password = st.sidebar.text_input("Enter Admin Password for Reports", type="password")
        
        if password == admin_password:
            view_reports_page()
        else:
            st.warning("Please enter the correct admin password to view reports.")
    elif page == "Admin Panel":
        admin_panel_page()

# Helper function to display the clock and date
def show_clock_and_date():
    # Ensure we always get the fresh system time
    today = datetime.now().strftime('%Y-%m-%d')
    current_time = datetime.now().strftime('%H:%M:%S')
    
    # First display: System time in Python
    st.markdown(f"""
        <h3 style="margin: 0; color: #0066cc;"> System Time :- {datetime.now().strftime('%H:%M')} </h3>
        <div style="font-size: 1.3rem; font-weight: bold;"></div>
        <div style="font-size: 1.3rem; font-weight: bold; margin-top: 5px;">{datetime.now().strftime('%A , %Y-%m-%d')}</div>
    </div>
    """, unsafe_allow_html=True)
    
    # Second display: Live JavaScript clock that updates in real-time
    st.markdown("""
    <script>
        // This function forces a new Date object creation on each call
        // to ensure we always get the current system time
        function updateClock() {
            // Get fresh system time
            var now = new Date();
            
            // Format date: Monday, May 7, 2025
            var options = { weekday: 'long', year: 'numeric', month: 'long', day: 'numeric' };
            var dateStr = now.toLocaleDateString(undefined, options);
            
            // Format time with hours, minutes, seconds
            var hours = now.getHours();
            var minutes = now.getMinutes();
            var seconds = now.getSeconds();
            
            // Add leading zeros
            hours = hours < 10 ? '0' + hours : hours;
            minutes = minutes < 10 ? '0' + minutes : minutes;
            seconds = seconds < 10 ? '0' + seconds : seconds;
            
            // 24-hour format time string
            var timeStr = hours + ':' + minutes + ':' + seconds;
            
            // Update the DOM elements
            document.getElementById('date_display').innerHTML = dateStr;
            document.getElementById('clock_display').innerHTML = timeStr;
            
            // Call again in exactly 1 second (1000ms)
            setTimeout(updateClock, 1000);
        }
        
        // Start the clock immediately when page loads
        updateClock();
    </script>
    """, unsafe_allow_html=True)
    
    # Note: We pass the current time from the server for operations that need it
    return today, current_time

# Helper function to display today's attendance dashboard
def show_attendance_dashboard():
    today = datetime.now().strftime('%Y-%m-%d')
    
    # Create a section with real-time attendance data
    st.subheader("Today's Attendance Dashboard (Real-Time)")
    
    # Create containers for auto-updating content
    attendance_container = st.container()
    
    with attendance_container:
        # Get fresh data every time
        df_latest = load_data()
        today_data = df_latest[df_latest['Date'] == today]
        
        if not today_data.empty:
            # Sort by Employee ID
            today_data = today_data.sort_values('Employee ID')
            
            # Display attendance statistics
            stat_col1, stat_col2, stat_col3, stat_col4 = st.columns(4)
            with stat_col1:
                st.metric("Total Employees", len(today_data['Employee ID'].unique()))
            with stat_col2:
                st.metric("Currently Punched In", len(today_data[today_data['Status'] == 'In Progress']))
            with stat_col3:
                st.metric("Completed Today", len(today_data[today_data['Status'] == 'Completed']))
            with stat_col4:
                if 'Is Late' in today_data.columns:
                    st.metric("Late Arrivals", len(today_data[today_data['Is Late'] == True]))
                else:
                    st.metric("Late Arrivals", "N/A")
            
            # Make it visually appealing with styled dataframe - red for late, green for on-time
            def style_dataframe(row):
                styles = []
                for _ in row:
                    if row['Status'] == 'Completed':
                        styles.append('background-color: #E0F7FA; color: black')  # Light blue for completed
                    elif 'Is Late' in row and row['Is Late']:
                        styles.append('background-color: #FFCCCC; color: black')  # Light red for late
                    else:
                        styles.append('background-color: #CCFFCC; color: black')  # Light green for on-time
                return styles
            
            st.dataframe(
                today_data.style.apply(style_dataframe, axis=1),
                use_container_width=True
            )
            
            # Show lateness statistics if we have that data
            if 'Is Late' in today_data.columns and not today_data.empty:
                st.subheader("Attendance Statistics")
                
                # Create pie chart data
                on_time_count = len(today_data[today_data['Is Late'] == False])
                late_count = len(today_data[today_data['Is Late'] == True])
                
                if on_time_count > 0 or late_count > 0:
                    # Create columns for side-by-side display
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        # Display metrics
                        st.metric("On-Time Arrivals", on_time_count)
                        st.metric("Late Arrivals", late_count)
                        
                        # Calculate punctuality rate
                        punctuality_rate = on_time_count / (on_time_count + late_count) * 100 if (on_time_count + late_count) > 0 else 0
                        st.metric("Punctuality Rate", f"{punctuality_rate:.1f}%")
                    
                    with col2:
                        # Create a simple pie chart using matplotlib
                        if on_time_count > 0 or late_count > 0:
                            fig, ax = plt.subplots(figsize=(4, 4))
                            ax.pie([on_time_count, late_count], 
                                  labels=['On Time', 'Late'], 
                                  autopct='%1.1f%%',
                                  colors=['#CCFFCC', '#FFCCCC'],
                                  startangle=90)
                            ax.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
                            plt.title('Punctuality Statistics')
                            st.pyplot(fig)
        else:
            st.info("No attendance records for today yet.")
            
    # Add automatic refresh message
    st.caption("Data refreshes automatically with page refresh")

# Punch In page
def punch_in_page():
    st.header("Vistotech Morning Punch In")
    
    # Get fresh current time directly from system
    current_time = datetime.now().strftime('%H:%M:%S')
    today = datetime.now().strftime('%Y-%m-%d')
    
    # Show clock interface
    show_clock_and_date()
    
    # Load registered employees data
    try:
        employees_df = load_employee_data()
        has_employee_registry = not employees_df.empty
    except:
        has_employee_registry = False
        employees_df = pd.DataFrame()
    
    # Input fields for employee information
    emp_id = st.text_input("Employee ID")
    
    if emp_id:
        # Check if this is a registered employee ID if we have employee data
        is_valid_employee = True
        if has_employee_registry:
            # Convert both to strings for comparison
            emp_id_str = str(emp_id).strip()
            registered_ids = employees_df['Employee ID'].astype(str).str.strip().tolist()
            
            if emp_id_str not in registered_ids:
                st.error(f"❌ Employee ID {emp_id} is not registered in the system. Please contact your administrator.")
                is_valid_employee = False
                
                # Show registered employees in an expander for admin reference
                with st.expander("Available Employee IDs"):
                    st.info("The following employee IDs are registered in the system:")
                    for idx, row in employees_df.iterrows():
                        st.write(f"- ID: {row['Employee ID']} | Name: {row['Employee Name']}")
                    st.caption("If you need to register a new employee, please use the Admin Panel.")
        
        if is_valid_employee:
            # Check if employee is already punched in - no name passed
            already_punched_in, index = check_existing_punch_in(emp_id, today) if emp_id else (False, None)
            
            # Load current record data
            df = load_data()
            
            if already_punched_in:
                st.warning(f"⚠️ Employee ID {emp_id} is already punched in for today. Please use the Punch Out option to complete your attendance.")
                
                # Show when they punched in
                current_record = df.iloc[index]
                st.info(f"You punched in at {current_record['Punch In Time']}")
                
                # Show a message directing them to the punch out option
                st.info("To punch out, please select the 'Punch Out' option from the sidebar menu.")
            else:
                # When not punched in, show punch in option
                st.write("### Record Morning Attendance")
                
                # Show punch in button with proper key to prevent button conflicts
                if st.button("📥 PUNCH IN", use_container_width=True, type="primary", key="main_punch_in"):
                    # Check for completed records for today
                    completed_today = df[(df['Employee ID'].astype(str).str.strip() == str(emp_id).strip()) & 
                                    (df['Date'] == today) & 
                                    (df['Status'] == 'Completed')]
                    
                    if not completed_today.empty:
                        st.error(f"You have already completed your attendance for today at {completed_today.iloc[0]['Punch Out Time']}.")
                    else:
                        # Store the punch in success in session state
                        if 'punch_in_success' not in st.session_state:
                            st.session_state.punch_in_success = False
                        
                        try:
                            # Check if punch in time is after 10:15 AM
                            is_late = False
                            current_time_obj = datetime.strptime(current_time, '%H:%M:%S').time()
                            cutoff_time = datetime.strptime("10:15:00", '%H:%M:%S').time()
                            
                            if current_time_obj > cutoff_time:
                                is_late = True
                                st.warning(f"⚠️ You are late! The cutoff time is 10:15 AM. You punched in at {current_time}.")
                            
                            # Get employee name from registry
                            emp_name = "Unknown"
                            if has_employee_registry:
                                emp_name = employees_df[employees_df['Employee ID'].astype(str).str.strip() == emp_id_str]['Employee Name'].iloc[0] if emp_id_str in registered_ids else "Unknown"
                            
                            # Create a new record
                            new_row = {
                                'Employee ID': emp_id,
                                'Employee Name': emp_name,
                                'Date': today,
                                'Punch In Time': current_time,
                                'Punch Out Time': None,
                                'Work Hours': None,
                                'Status': 'In Progress',
                                'Is Late': is_late
                            }
                            
                            # Add the new row
                            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
                            
                            # Save data to Excel
                            if save_data(df):
                                # Apply color formatting
                                apply_excel_formatting()
                                
                                st.session_state.punch_in_success = True
                                
                                if is_late:
                                    st.warning(f"⚠️ Late Punch In recorded at {current_time}. Your entry has been marked as LATE.")
                                else:
                                    st.success(f"✅ On-time Punch In recorded at {current_time}.")
                                    
                                st.success(f"Employee ID: {emp_id} successfully punched in!")
                                st.balloons()
                                
                        except Exception as e:
                            st.error(f"Error saving punch in record: {e}")
                            st.info("Please try again.")
                        
        # Show status in a separate section
        st.markdown("---")
        
        # Show the current status for this employee
        st.subheader(f"Current Status for Employee ID {emp_id}")
        
        # Check current status again to ensure we have the latest data
        already_punched_in, index = check_existing_punch_in(emp_id, today)
        
        if already_punched_in:
            current_record = df.iloc[index]
            # Check if the employee was late
            if 'Is Late' in current_record and current_record['Is Late']:
                st.warning(f"📌 Status: Employee ID {emp_id} is currently PUNCHED IN (LATE at {current_record['Punch In Time']})")
            else:
                st.success(f"📌 Status: Employee ID {emp_id} is currently PUNCHED IN (ON TIME at {current_record['Punch In Time']})")
        else:
            completed_today = df[(df['Employee ID'] == emp_id) & 
                              (df['Date'] == today) & 
                              (df['Status'] == 'Completed')]
            
            if not completed_today.empty:
                punch_in = completed_today.iloc[0]['Punch In Time']
                punch_out = completed_today.iloc[0]['Punch Out Time'] 
                work_hours = completed_today.iloc[0]['Work Hours']
                is_late = completed_today.iloc[0]['Is Late'] if 'Is Late' in completed_today.iloc[0] else False
                
                if is_late:
                    st.info(f"📌 Status: Employee ID {emp_id} has COMPLETED attendance for today (LATE)")
                else:
                    st.info(f"📌 Status: Employee ID {emp_id} has COMPLETED attendance for today (ON TIME)")
                
                st.success(f"Punch In: {punch_in} | Punch Out: {punch_out} | Work Hours: {work_hours} hrs")
            else:
                st.warning(f"📌 Status: Employee ID {emp_id} is NOT punched in")
    else:
        st.info("Please enter your Employee ID to punch in.")
    
    # Show the attendance dashboard
    show_attendance_dashboard()

# Punch Out page
def punch_out_page():
    st.header("Vistotech Evening Punch Out")
    
    # Get fresh current time directly from system
    current_time = datetime.now().strftime('%H:%M:%S')
    today = datetime.now().strftime('%Y-%m-%d')
    
    # Show clock interface
    show_clock_and_date()
    
    # Load registered employees data
    try:
        employees_df = load_employee_data()
        has_employee_registry = not employees_df.empty
    except:
        has_employee_registry = False
        employees_df = pd.DataFrame()
        
    # Input fields for employee information
    emp_id = st.text_input("Employee ID")
    
    if emp_id:
        # Check if this is a registered employee ID if we have employee data
        is_valid_employee = True
        if has_employee_registry:
            # Convert both to strings for comparison
            emp_id_str = str(emp_id).strip()
            registered_ids = employees_df['Employee ID'].astype(str).str.strip().tolist()
            
            if emp_id_str not in registered_ids:
                st.error(f"❌ Employee ID {emp_id} is not registered in the system. Please contact your administrator.")
                is_valid_employee = False
                
                # Show registered employees in an expander for admin reference
                with st.expander("Available Employee IDs"):
                    st.info("The following employee IDs are registered in the system:")
                    for idx, row in employees_df.iterrows():
                        st.write(f"- ID: {row['Employee ID']} | Name: {row['Employee Name']}")
                    st.caption("If you need to register a new employee, please use the Admin Panel.")
        
        if is_valid_employee:
            # Load data first
            df = load_data()
            
            # Convert the emp_id to string to ensure consistent comparison
            emp_id_str = str(emp_id).strip()
            
            # Get all records for this employee ID today
            today_records = df[df['Date'] == today]
            today_records = today_records[today_records['Employee ID'].astype(str).str.strip() == emp_id_str]
            
            # Check if there are any records for this employee ID
            if today_records.empty:
                st.error(f"No attendance records found for Employee ID: {emp_id} today. Please punch in first.")
                return
                
            # First try checking for in-progress records
            in_progress_records = today_records[today_records['Status'] == 'In Progress']
            already_punched_in = not in_progress_records.empty
            
            if already_punched_in:
                # If we found in-progress records, get the first one
                index = in_progress_records.index[0]
                st.success(f"Found punch-in record for Employee ID: {emp_id}!")
            else:
                # If not found, show available records for this ID and ask for confirmation
                st.warning(f"No in-progress record found for Employee ID {emp_id}.")
                
                # Check if they've completed records for today
                completed_records = today_records[today_records['Status'] == 'Completed']
                if not completed_records.empty:
                    st.info("You have already completed your attendance for today:")
                    for _, record in completed_records.iterrows():
                        st.success(f"Punch In: {record['Punch In Time']} | Punch Out: {record['Punch Out Time']} | Work Hours: {record['Work Hours']} hrs")
                    return
                else:
                    st.error("No valid records found. Please punch in first using the 'Punch In' option.")
                    return
        
        # If we found a record to punch out
        if is_valid_employee and already_punched_in:
            current_record = df.iloc[index]
            
            # When already punched in, show punch out option
            st.write("### Record End of Day Punch Out")
            
            # Show when they punched in
            punch_in_time = current_record['Punch In Time']
            is_late = current_record['Is Late'] if 'Is Late' in current_record else False
            
            if is_late:
                st.warning(f"You punched in LATE at {punch_in_time}")
            else:
                st.success(f"You punched in ON TIME at {punch_in_time}")
            
            # Store the punch out action in a session state
            if 'punch_out_success' not in st.session_state:
                st.session_state.punch_out_success = False
            
            # Show punch out button
            if st.button("📤 PUNCH OUT", use_container_width=True, type="primary"):
                if current_record is None:
                    st.error("Error accessing your punch record. Please try again.")
                else:
                    # Update punch out time
                    # Calculate work hours
                    work_hours = calculate_hours(current_record['Punch In Time'], current_time)
                    
                    df.at[index, 'Punch Out Time'] = current_time
                    df.at[index, 'Work Hours'] = work_hours
                    df.at[index, 'Status'] = 'Completed'
                    
                    if save_data(df):
                        st.session_state.punch_out_success = True
                        st.success(f"✅ Punch Out recorded at {current_time} for Employee ID {emp_id}")
                        st.success(f"Total work hours for today: {work_hours} hrs")
                        st.balloons()
            
            # Display success message if previously punched out
            if st.session_state.punch_out_success:
                st.success("You have successfully punched out for today!")
                
        elif is_valid_employee:
            # Check if they've already completed attendance for today
            completed_today = df[(df['Employee ID'].astype(str).str.strip() == str(emp_id).strip()) & 
                             (df['Date'] == today) & 
                             (df['Status'] == 'Completed')]
            
            if not completed_today.empty:
                # They've already punched out
                punch_in = completed_today.iloc[0]['Punch In Time']
                punch_out = completed_today.iloc[0]['Punch Out Time']
                work_hours = completed_today.iloc[0]['Work Hours']
                
                st.info("You have already completed your attendance for today:")
                st.success(f"Punch In: {punch_in} | Punch Out: {punch_out} | Work Hours: {work_hours} hrs")
            else:
                # They haven't punched in yet
                st.warning(f"⚠️ Employee ID {emp_id} has not punched in for today.")
                st.info("Please use the 'Punch In' option from the sidebar menu to record your entry first.")
        
        # Show status in a separate section
        st.markdown("---")
        
        # Show the current status for this employee
        st.subheader(f"Current Status for Employee ID {emp_id}")
        
        # Check current status again to ensure we have the latest data
        already_punched_in, index = check_existing_punch_in(emp_id, today)
        
        if already_punched_in:
            current_record = df.iloc[index]
            # Check if the employee was late
            if 'Is Late' in current_record and current_record['Is Late']:
                st.warning(f"📌 Status: Employee ID {emp_id} is currently PUNCHED IN (LATE at {current_record['Punch In Time']})")
            else:
                st.success(f"📌 Status: Employee ID {emp_id} is currently PUNCHED IN (ON TIME at {current_record['Punch In Time']})")
        else:
            completed_today = df[(df['Employee ID'] == emp_id) & 
                              (df['Date'] == today) & 
                              (df['Status'] == 'Completed')]
            
            if not completed_today.empty:
                punch_in = completed_today.iloc[0]['Punch In Time']
                punch_out = completed_today.iloc[0]['Punch Out Time'] 
                work_hours = completed_today.iloc[0]['Work Hours']
                is_late = completed_today.iloc[0]['Is Late'] if 'Is Late' in completed_today.iloc[0] else False
                
                if is_late:
                    st.info(f"📌 Status: Employee ID {emp_id} has COMPLETED attendance for today (LATE)")
                else:
                    st.info(f"📌 Status: Employee ID {emp_id} has COMPLETED attendance for today (ON TIME)")
                
                st.success(f"Punch In: {punch_in} | Punch Out: {punch_out} | Work Hours: {work_hours} hrs")
            else:
                st.warning(f"📌 Status: Employee ID {emp_id} is NOT punched in")
    else:
        st.info("Please enter your Employee ID to punch out.")
    
    # Show the attendance dashboard
    show_attendance_dashboard()

# View Reports page
def view_reports_page():
    st.header("Vistotech Attendance Reports")
    
    # Load the data
    df = load_data()
    
    if df.empty:
        st.info("No attendance data available.")
        return
    
    # Date range selection
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("Start Date", 
                                 min_value=pd.to_datetime(df['Date']).dt.date.min() if not df.empty else datetime.now().date())
    with col2:
        end_date = st.date_input("End Date", 
                               max_value=pd.to_datetime(df['Date']).dt.date.max() if not df.empty else datetime.now().date())
    
    # Convert to string for filtering
    start_date_str = start_date.strftime('%Y-%m-%d')
    end_date_str = end_date.strftime('%Y-%m-%d')
    
    # Filter data by date range
    filtered_df = df[(df['Date'] >= start_date_str) & (df['Date'] <= end_date_str)]
    
    # Employee filter
    if not df['Employee ID'].empty:
        employee_filter = st.multiselect("Filter by Employee", 
                                      options=sorted(df['Employee ID'].unique()), 
                                      default=[])
        
        if employee_filter:
            filtered_df = filtered_df[filtered_df['Employee ID'].isin(employee_filter)]
    
    # Add lateness filter
    if 'Is Late' in filtered_df.columns:
        lateness_filter = st.multiselect("Filter by Punctuality",
                                       options=["On Time", "Late"],
                                       default=[])
        
        if lateness_filter:
            if "On Time" in lateness_filter and "Late" in lateness_filter:
                pass  # Show all records
            elif "On Time" in lateness_filter:
                filtered_df = filtered_df[filtered_df['Is Late'] == False]
            elif "Late" in lateness_filter:
                filtered_df = filtered_df[filtered_df['Is Late'] == True]
    
    # Show filtered data
    if not filtered_df.empty:
        st.subheader("Filtered Attendance Records")
        
        # Style the dataframe with colors
        def style_dataframe(row):
            styles = []
            for _ in row:
                if row['Status'] == 'Completed':
                    styles.append('background-color: #E0F7FA; color: black')  # Light blue for completed
                elif 'Is Late' in row and row['Is Late']:
                    styles.append('background-color: #FFCCCC; color: black')  # Light red for late
                else:
                    styles.append('background-color: #CCFFCC; color: black')  # Light green for on-time
            return styles
        
        st.dataframe(
            filtered_df.style.apply(style_dataframe, axis=1),
            use_container_width=True
        )
        
        # Calculate total hours worked and lateness statistics
        if 'Work Hours' in filtered_df.columns and not filtered_df[filtered_df['Work Hours'].notna()].empty:
            # Calculate hours by employee
            hours_by_employee = filtered_df.groupby(['Employee ID', 'Employee Name'])['Work Hours'].sum().reset_index()
            
            # Calculate lateness count by employee if we have that data
            if 'Is Late' in filtered_df.columns:
                lateness_by_employee = filtered_df.groupby(['Employee ID', 'Employee Name'])['Is Late'].sum().reset_index()
                lateness_by_employee = lateness_by_employee.rename(columns={'Is Late': 'Late Count'})
                
                # Count days present
                days_present = filtered_df.groupby(['Employee ID', 'Employee Name']).size().reset_index(name='Days Present')
                
                # Merge the data
                summary_df = pd.merge(hours_by_employee, lateness_by_employee, on=['Employee ID', 'Employee Name'])
                summary_df = pd.merge(summary_df, days_present, on=['Employee ID', 'Employee Name'])
                
                # Calculate punctuality rate
                summary_df['Punctuality Rate'] = ((summary_df['Days Present'] - summary_df['Late Count']) / summary_df['Days Present'] * 100).round(1)
                
                # Show summary
                st.subheader("Employee Summary")
                st.dataframe(summary_df, use_container_width=True)
            else:
                # Just show hours summary if we don't have lateness data
                st.subheader("Total Hours Summary")
                st.dataframe(hours_by_employee, use_container_width=True)
            
            # Show visualization of hours worked
            st.subheader("Hours Worked by Employee")
            st.bar_chart(hours_by_employee.set_index('Employee Name')['Work Hours'])
            
            # Show lateness visualization if we have that data
            if 'Is Late' in filtered_df.columns:
                late_counts = filtered_df.groupby('Date')['Is Late'].sum().reset_index()
                late_counts['Date'] = pd.to_datetime(late_counts['Date'])
                late_counts = late_counts.sort_values('Date')
                
                st.subheader("Late Arrivals by Date")
                st.line_chart(late_counts.set_index('Date')['Is Late'])
    else:
        st.info("No records found for the selected criteria.")
    
    # Export functionality
    if not filtered_df.empty:
        st.download_button(
            label="Export to CSV",
            data=filtered_df.to_csv(index=False).encode('utf-8'),
            file_name=f"attendance_report_{start_date_str}_to_{end_date_str}.csv",
            mime="text/csv"
        )

# Helper function to check password strength
def check_password_strength(password):
    """
    Check if password meets strength requirements:
    - At least 8 characters
    - Contains uppercase letter
    - Contains lowercase letter
    - Contains number
    - Contains special character (@, $, etc.)
    
    Returns: (bool, str) - (is_strong, message)
    """
    errors = []
    
    if len(password) < 8:
        errors.append("Password must be at least 8 characters long")
    
    if not any(c.isupper() for c in password):
        errors.append("Password must contain at least one uppercase letter")
    
    if not any(c.islower() for c in password):
        errors.append("Password must contain at least one lowercase letter")
    
    if not any(c.isdigit() for c in password):
        errors.append("Password must contain at least one number")
    
    if not any(c in "@$!%*?&" for c in password):
        errors.append("Password must contain at least one special character (@, $, !, %, *, ?, &)")
    
    is_strong = len(errors) == 0
    message = "\n".join(errors) if errors else "Password meets strength requirements"
    
    return is_strong, message

# Function to save admin password
def save_admin_password(new_password):
    """Save admin password to a secure file"""
    # In a real application, you would hash the password and store it securely
    # For this demo, we'll use a simple text file
    with open("admin_password.txt", "w") as f:
        f.write(new_password)
    return True

# Function to get admin password
def get_admin_password():
    """Get admin password from file or return default"""
    try:
        with open("admin_password.txt", "r") as f:
            return f.read().strip()
    except FileNotFoundError:
        # Default password if file doesn't exist
        return "admin123"

# Function to load and save employee records
def load_employee_data():
    """Load employee data from the EMPLOYEES_FILE"""
    EMPLOYEES_FILE = "employees.xlsx"
    try:
        if os.path.exists(EMPLOYEES_FILE):
            df = pd.read_excel(EMPLOYEES_FILE)
            return df
        else:
            # Create a new DataFrame if file doesn't exist
            df = pd.DataFrame(columns=['Employee ID', 'Employee Name', 'Date Added'])
            df.to_excel(EMPLOYEES_FILE, index=False)
            return df
    except Exception as e:
        st.error(f"Error loading employee data: {e}")
        return pd.DataFrame(columns=['Employee ID', 'Employee Name', 'Date Added'])

def save_employee_data(df):
    """Save employee data to EMPLOYEES_FILE"""
    EMPLOYEES_FILE = "employees.xlsx"
    try:
        df.to_excel(EMPLOYEES_FILE, index=False)
        return True
    except Exception as e:
        st.error(f"Error saving employee data: {e}")
        return False

# Admin Panel page
def admin_panel_page():
    st.header("Vistotech Admin Panel")
    
    # Always get fresh time
    current_time = datetime.now().strftime('%H:%M:%S')
    today = datetime.now().strftime('%Y-%m-%d')
    
    # Display current system time at the top of admin panel for reference
    st.markdown(f"""
        <div style="font-weight: bold; color: #0066cc;">Current System Time: {datetime.now().strftime('%H:%M:%S')}</div>
        <div style="font-size: 1.2rem; color: #0066cc;">{datetime.now().strftime('%A %Y-%m-%d')}</div>
    </div>
    """, unsafe_allow_html=True)
    
    # Load data
    df = load_data()
    
    # Add password protection
    current_admin_password = get_admin_password()
    password = st.text_input("Enter Admin Password", type="password")
    
    if password != current_admin_password:
        st.warning("Please enter the correct password to access admin features.")
        return
    
    st.success("Admin access granted!")
    
    # Admin actions tabs
    admin_tab1, admin_tab2, admin_tab3, admin_tab4 = st.tabs([
        "Attendance Records", 
        "Employee Management", 
        "Change Password", 
        "System Settings"
    ])
    
    # Tab 1: Attendance Records
    with admin_tab1:
        st.subheader("All Attendance Records")
        if not df.empty:
            # Apply styling to the dataframe
            def style_dataframe(row):
                styles = []
                for _ in row:
                    if row['Status'] == 'Completed':
                        styles.append('background-color: #E0F7FA; color: black')  # Light blue for completed
                    elif 'Is Late' in row and row['Is Late']:
                        styles.append('background-color: #FFCCCC; color: black')  # Light red for late
                    else:
                        styles.append('background-color: #CCFFCC; color: #000080')  # Light green for on-time
                return styles
            
            st.dataframe(
                df.style.apply(style_dataframe, axis=1),
                use_container_width=True
            )
            
            # Manual record editing
            with st.expander("Edit Records"):
                # Select record to edit
                selected_index = st.selectbox("Select record to edit", 
                                            options=df.index,
                                            format_func=lambda x: f"{df.loc[x, 'Employee ID']} - {df.loc[x, 'Employee Name']} - {df.loc[x, 'Date']}")
                
                if selected_index is not None:
                    # Create a 2-column layout for time fields
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        new_punch_in = st.text_input("Punch In Time", value=df.loc[selected_index, 'Punch In Time'])
                        
                        # Determine if this punch in would be considered late
                        try:
                            punch_in_time = datetime.strptime(new_punch_in, '%H:%M:%S').time()
                            cutoff_time = datetime.strptime("10:15:00", '%H:%M:%S').time()
                            is_late = punch_in_time > cutoff_time
                            
                            if is_late:
                                st.warning("⚠️ This punch in time is after 10:15 AM and will be marked as LATE")
                            else:
                                st.success("✅ This punch in time is before 10:15 AM and will be marked as ON TIME")
                        except:
                            pass
                        
                    with col2:
                        new_punch_out = st.text_input("Punch Out Time", 
                                                    value=df.loc[selected_index, 'Punch Out Time'] 
                                                    if pd.notna(df.loc[selected_index, 'Punch Out Time']) else "")
                    
                    # Update button
                    if st.button("Update Record"):
                        # Update punch in time and check if it's late
                        df.at[selected_index, 'Punch In Time'] = new_punch_in
                        
                        try:
                            punch_in_time = datetime.strptime(new_punch_in, '%H:%M:%S').time()
                            cutoff_time = datetime.strptime("10:15:00", '%H:%M:%S').time()
                            df.at[selected_index, 'Is Late'] = punch_in_time > cutoff_time
                        except:
                            pass
                        
                        # Update punch out time if provided
                        if new_punch_out:
                            df.at[selected_index, 'Punch Out Time'] = new_punch_out
                            # Recalculate work hours
                            work_hours = calculate_hours(new_punch_in, new_punch_out)
                            df.at[selected_index, 'Work Hours'] = work_hours
                            df.at[selected_index, 'Status'] = 'Completed'
                        
                        # Save data and apply formatting
                        if save_data(df):
                            apply_excel_formatting()
                            st.success("✅ Record updated successfully")
                            st.rerun()  # Refresh the page to show updates
            
            # Clear All Records button with confirmation
            with st.expander("Clear Records"):
                st.warning("⚠️ This will permanently delete all attendance records. This action cannot be undone.")
                confirm_delete = st.text_input("Type 'DELETE' to confirm clearing all records", key="confirm_delete")
                
                if confirm_delete == "DELETE":
                    if st.button("Permanently Clear All Records", type="primary"):
                        # Create a new empty DataFrame with the same columns
                        new_df = pd.DataFrame(columns=df.columns)
                        if save_data(new_df):
                            st.success("✅ All records have been cleared")
                            time.sleep(1)
                            st.rerun()
        else:
            st.info("No attendance records available.")
    
    # Tab 2: Employee Management
    with admin_tab2:
        st.subheader("Employee Management")
        
        # Load employee data
        employees_df = load_employee_data()
        
        # Display existing employees
        if not employees_df.empty:
            st.write("**Registered Employees:**")
            st.dataframe(employees_df, use_container_width=True)
            
            # Count total employees
            st.info(f"Total registered employees: {len(employees_df)}")
        else:
            st.info("No employees registered yet.")
        
        # Add new employee section
        st.markdown("---")
        st.subheader("Add New Employee")
        
        # Input fields for new employee
        new_emp_id = st.text_input("New Employee ID", placeholder="e.g., 1001")
        new_emp_name = st.text_input("New Employee Name", placeholder="e.g., John Doe")
        
        # Add employee button
        if new_emp_id and new_emp_name:
            # Check if ID already exists
            if not employees_df.empty and new_emp_id in employees_df['Employee ID'].astype(str).values:
                st.error(f"❌ Employee ID {new_emp_id} already exists! Please use a unique ID.")
            else:
                if st.button("Add Employee", type="primary"):
                    # Add new employee to dataframe
                    new_row = {
                        'Employee ID': new_emp_id,
                        'Employee Name': new_emp_name,
                        'Date Added': datetime.now().strftime('%Y-%m-%d')
                    }
                    
                    employees_df = pd.concat([employees_df, pd.DataFrame([new_row])], ignore_index=True)
                    
                    # Save updated employee data
                    if save_employee_data(employees_df):
                        st.success(f"✅ Employee {new_emp_name} (ID: {new_emp_id}) added successfully!")
                        st.rerun()  # Refresh the page to show new employee
                    else:
                        st.error("Failed to save employee data.")
        
        # Delete employee section
        st.markdown("---")
        st.subheader("Delete Employee")
        
        if not employees_df.empty:
            # Select employee to delete
            delete_emp_index = st.selectbox(
                "Select employee to delete", 
                options=employees_df.index,
                format_func=lambda x: f"{employees_df.iloc[x]['Employee ID']} - {employees_df.iloc[x]['Employee Name']}"
            )
            
            if delete_emp_index is not None:
                emp_to_delete = employees_df.iloc[delete_emp_index]
                st.warning(f"⚠️ About to delete: {emp_to_delete['Employee Name']} (ID: {emp_to_delete['Employee ID']})")
                
                # Confirmation for deletion
                confirm_emp_delete = st.text_input("Type 'DELETE' to confirm", key="confirm_emp_delete")
                
                if confirm_emp_delete == "DELETE":
                    if st.button("Delete Employee", key="delete_emp_button", type="primary"):
                        # Remove employee from dataframe
                        employees_df = employees_df.drop(delete_emp_index).reset_index(drop=True)
                        
                        # Save updated employee data
                        if save_employee_data(employees_df):
                            st.success(f"✅ Employee deleted successfully!")
                            st.rerun()  # Refresh the page
                        else:
                            st.error("Failed to save employee data.")
    
    # Tab 3: Change Password
    with admin_tab3:
        st.subheader("Change Admin Password")
        
        new_password = st.text_input("New Password", type="password")
        confirm_password = st.text_input("Confirm New Password", type="password")
        
        if new_password and confirm_password:
            if new_password != confirm_password:
                st.error("Passwords do not match.")
            else:
                # Check password strength
                is_strong, message = check_password_strength(new_password)
                
                if not is_strong:
                    st.error(message)
                else:
                    st.success(message)
                    
                    if st.button("Change Password"):
                        if save_admin_password(new_password):
                            st.success("✅ Password changed successfully")
                            
                            # Force refresh
                            time.sleep(1)
                            st.rerun()
    
    # Tab 4: System Settings
    with admin_tab4:
        st.subheader("System Settings")
        
        # Attendance cutoff time setting
        st.info("Attendance Settings")
        st.write("Current punctuality cutoff time: **10:15 AM**")
        st.write("Employees who punch in after this time will be marked as late.")
        
        # Excel Formatting
        st.subheader("Excel Formatting")
        if st.button("Re-apply Excel Color Formatting"):
            if apply_excel_formatting():
                st.success("✅ Excel formatting has been applied successfully")
            else:
                st.error("❌ Error applying Excel formatting")
        
        # About & Information
        st.subheader("System Information")
        st.write("Vistotech Attendance System v1.0")
        st.write("Date: May 2025")
        st.write("Total records in database:", len(df) if not df.empty else 0)
        st.write("Total registered employees:", len(load_employee_data()))

# Run the app
if __name__ == "__main__":
    main()